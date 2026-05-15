# ekstrak_kbbi.py
# Script untuk mengekstrak data dari kbbi_v(rapih).xlsx
# Menghasilkan 3 file CSV terpisah untuk 3 algoritma berbeda

import openpyxl
import re
import os

def bersihkan_teks(teks):
    """Membersihkan teks dari karakter yang tidak perlu."""
    if not teks or str(teks).strip() == '' or str(teks).lower() == 'none':
        return None
    teks = str(teks).strip().lower()
    # Buang sampah dari Excel (formula objects, dll)
    if 'openpyxl' in teks or '<' in teks or '>' in teks:
        return None
    # Hapus teks dalam kurung seperti (1), (2), (Mk), dll
    teks = re.sub(r'\(\d+\)', '', teks)
    teks = re.sub(r'\([A-Za-z]{1,3}\)', '', teks)
    # Hapus titik pelafalan KBBI (contoh: "i.ni" -> "ini", "me.la.ku.kan" -> "melakukan")
    # Tapi jangan hapus titik kalau itu singkatan (ada huruf besar) atau angka
    teks = re.sub(r'(?<=[a-z])\.(?=[a-z])', '', teks)
    teks = teks.strip('. ')
    return teks if len(teks) > 0 else None

def adalah_kata_tunggal(kata):
    """Cek apakah ini kata tunggal (tanpa spasi) dan layak masuk kamus."""
    if not kata:
        return False
    kata = kata.strip()
    # Harus lebih dari 1 karakter
    if len(kata) <= 1:
        return False
    # Tidak boleh mengandung spasi (itu frasa, bukan kata tunggal)
    if ' ' in kata:
        return False
    # Hanya boleh huruf dan tanda hubung (untuk kata seperti "ber-", "se-")
    if not re.match(r'^[a-z\-]+$', kata):
        return False
    # Buang yang hanya berisi tanda hubung
    if kata.replace('-', '') == '':
        return False
    return True

def ekstrak_frasa(teks):
    """Mengekstrak frasa gabungan kata (yang mengandung spasi)."""
    if not teks:
        return []
    frasa_list = []
    # Gabungan kata sering dipisah koma atau titik koma
    potongan = re.split(r'[;,]', str(teks))
    for p in potongan:
        p = bersihkan_teks(p)
        if p and ' ' in p:
            # Bersihkan karakter non-huruf kecuali spasi dan hubung
            p = re.sub(r'[^a-z\s\-]', '', p).strip()
            p = re.sub(r'\s+', ' ', p)  # Normalisasi spasi ganda
            if p and len(p) > 3 and ' ' in p:
                frasa_list.append(p)
    return frasa_list

def main():
    print("[START] Memulai ekstraksi data KBBI...")
    print("   Membuka file kbbi_v(rapih).xlsx...")
    
    wb = openpyxl.load_workbook('kbbi_v(rapih).xlsx', read_only=True)
    ws = wb.active
    
    # Ambil header
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"   Kolom ditemukan: {headers}")
    
    # Siapkan wadah data
    kamus_kata = set()          # Untuk Levenshtein (kata tunggal)
    kamus_frasa = set()         # Untuk N-Gram (frasa berspasi)
    kamus_kelas = {}            # Untuk POS Tagging (kata → kelas)
    kamus_tidak_baku = {}       # Bonus: bentuk tidak baku → bentuk baku
    
    # Cari index kolom yang dibutuhkan
    col_idx = {}
    for i, h in enumerate(headers):
        if h:
            col_idx[h] = i
    
    total_rows = 0
    
    print("   Memproses baris data...")
    for row in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        if total_rows % 20000 == 0:
            print(f"   ... sudah {total_rows} baris diproses")
        
        # --- KOLOM: nama ---
        nama = bersihkan_teks(row[col_idx.get('nama', 1)])
        if nama:
            if adalah_kata_tunggal(nama):
                kamus_kata.add(nama)
            elif ' ' in nama:
                kamus_frasa.add(nama)
        
        # --- KOLOM: kata_dasar ---
        kata_dasar = row[col_idx.get('kata_dasar', 3)]
        if kata_dasar and str(kata_dasar).lower() != 'none':
            # Bisa berisi beberapa kata dasar dipisah koma
            for kd in str(kata_dasar).split(','):
                kd = bersihkan_teks(kd)
                if kd and adalah_kata_tunggal(kd):
                    kamus_kata.add(kd)
        
        # --- KOLOM: kata_turunan ---
        kata_turunan = row[col_idx.get('kata_turunan', 12)]
        if kata_turunan and str(kata_turunan).lower() != 'none':
            for kt in re.split(r'[;,]', str(kata_turunan)):
                kt = bersihkan_teks(kt)
                if kt and adalah_kata_tunggal(kt):
                    kamus_kata.add(kt)
        
        # --- KOLOM: gabungan_kata (PENTING untuk N-Gram) ---
        gabungan = row[col_idx.get('gabungan_kata', 13)]
        if gabungan and str(gabungan).lower() != 'none':
            frasa_list = ekstrak_frasa(gabungan)
            for f in frasa_list:
                kamus_frasa.add(f)
        
        # --- KOLOM: kelas (untuk POS Tagging) ---
        kelas = row[col_idx.get('kelas', 7)]
        if nama and kelas and str(kelas).lower() != 'none' and str(kelas).strip() != '-':
            kata_kunci = nama if adalah_kata_tunggal(nama) else None
            if kata_kunci:
                kamus_kelas[kata_kunci] = str(kelas).strip()
        
        # --- KOLOM: bentuk_tidak_baku (BONUS) ---
        tidak_baku = row[col_idx.get('bentuk_tidak_baku', 5)]
        if nama and tidak_baku and str(tidak_baku).lower() != 'none':
            for tb in re.split(r'[;,]', str(tidak_baku)):
                tb = bersihkan_teks(tb)
                if tb and len(tb) > 1:
                    # tidak_baku → nama (bentuk baku)
                    bentuk_baku = nama if adalah_kata_tunggal(nama) else None
                    if bentuk_baku:
                        kamus_tidak_baku[tb] = bentuk_baku
    
    wb.close()
    
    # ===== SIMPAN HASIL =====
    output_dir = os.path.join('nlp_engine', 'data')
    os.makedirs(output_dir, exist_ok=True)
    
    # 1. Kamus kata tunggal (untuk Levenshtein)
    path_kata = os.path.join(output_dir, 'kamus_kata_baku.csv')
    with open(path_kata, 'w', encoding='utf-8') as f:
        f.write('kata_baku\n')
        for kata in sorted(kamus_kata):
            f.write(f'{kata}\n')
    
    # 2. Kamus frasa (untuk N-Gram)
    path_frasa = os.path.join(output_dir, 'kamus_frasa.csv')
    with open(path_frasa, 'w', encoding='utf-8') as f:
        f.write('frasa\n')
        for frasa in sorted(kamus_frasa):
            f.write(f'{frasa}\n')
    
    # 3. Kamus kelas kata (untuk POS Tagging)
    path_kelas = os.path.join(output_dir, 'kamus_kelas_kata.csv')
    with open(path_kelas, 'w', encoding='utf-8') as f:
        f.write('kata,kelas\n')
        for kata in sorted(kamus_kelas.keys()):
            kelas = kamus_kelas[kata].replace(',', ';')  # Escape koma
            f.write(f'{kata},{kelas}\n')
    
    # 4. Kamus bentuk tidak baku (BONUS)
    path_tidak_baku = os.path.join(output_dir, 'kamus_tidak_baku.csv')
    with open(path_tidak_baku, 'w', encoding='utf-8') as f:
        f.write('tidak_baku,bentuk_baku\n')
        for tb in sorted(kamus_tidak_baku.keys()):
            f.write(f'{tb},{kamus_tidak_baku[tb]}\n')
    
    # ===== LAPORAN =====
    print(f"\n{'='*50}")
    print(f"[DONE] EKSTRAKSI SELESAI!")
    print(f"{'='*50}")
    print(f"   Total baris diproses  : {total_rows:,}")
    print(f"   Kata tunggal (Levenshtein) : {len(kamus_kata):,} kata")
    print(f"   Frasa (N-Gram)             : {len(kamus_frasa):,} frasa")
    print(f"   Kata + Kelas (POS Tagging) : {len(kamus_kelas):,} entri")
    print(f"   Bentuk tidak baku (Bonus)  : {len(kamus_tidak_baku):,} entri")
    print(f"\n[OUTPUT] File tersimpan di folder: {output_dir}/")
    print(f"   +-- kamus_kata_baku.csv")
    print(f"   +-- kamus_frasa.csv")
    print(f"   +-- kamus_kelas_kata.csv")
    print(f"   +-- kamus_tidak_baku.csv")

if __name__ == '__main__':
    main()
